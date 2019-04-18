import os
import datetime
import glob
from time import sleep
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
from colour import Color
import openpyxl
import random

title = u'Title of poster'
subtitle = u'Subtitle'
buttonText = u'DOWNLOAD'

buttonHeight = 40
buttonWidth = 280

fillColor = (255,255,255)
outlineColor = (255,255,255)

x0=150
x1=x0+buttonWidth
y0=400
y1=y0+buttonHeight

category = '电子数码'

wb = openpyxl.load_workbook('/Users/moqu/Downloads/素材库1.0/需求/素材需求0410.xlsx')
sheet = wb['Sheet1']

def watermark():
    i=2
    for files in glob.glob('/Users/moqu/Downloads/imglib/*.png'):
        chfont = ImageFont.truetype('/Library/Fonts/Tahoma.ttf',72)
        enfont = ImageFont.truetype('/Library/Fonts/Tahoma.ttf',48)
        buttonfont = ImageFont.truetype('/Library/Fonts/Tahoma.ttf',32)
        img = Image.open(files)

        result = sheet.cell(row=i, column=6).value
        result_list = result.split("\n")
        title = str(result_list[0])
        subtitle = str(result_list[1])
        buttonText = str(result_list[2])
        language = sheet.cell(row=i,column=1).value
        elements = sheet.cell(row=i,column=10).value
        elementsfolder = '/Users/moqu/Desktop/广告平台/垃圾堆/商品1.1/'+elements+'/'
        ls = []
        for dirpath, dirnames, filenames in os.walk(elementsfolder):
            for filepath in filenames:
                ls.append(elementsfolder+filepath)
        #print(ls)
        rs = random.sample(range(1,len(ls)-1),3)
        for cornermark in rs:
            commodity_single_img = ls[cornermark]
            print (commodity_single_img)
            commodity_single_img = Image.open(commodity_single_img)
            commodity_region = commodity_single_img
            commodity_region = commodity_region.convert("RGBA")
            commodity_region.thumbnail((300, 200))
            img.paste(commodity_region, (800+random.randint(-100,100), 200+random.randint(-50,50)), commodity_region)

        #print(elementsfolder)

        draw = ImageDraw.Draw(img)
        draw.text((img.size[0] - 1060, img.size[1] - 416), title, fill=(255, 100, 100), font=chfont)
        draw.text((img.size[0] - 1060, img.size[1] - 420), title, fill=(255, 255, 255), font=chfont)
        draw.text((img.size[0] - 1060, img.size[1] - 320), subtitle, fill=(255, 255, 255), font=enfont)

        if(language == 'TH'):# Mid Width
            button_img = Image.open('/Users/moqu/Downloads/素材库1.0/按钮/RoundedRect.png').convert("RGBA")
            button_img.resize((1600,120)) # Button大小
            img.paste(button_img,(160,400),button_img)

        elif(language == 'PH'):
            button_img = Image.open('/Users/moqu/Downloads/素材库1.0/按钮/XLargeRoundedRect.png').convert("RGBA")
            button_img.resize((1600,120)) # Button大小
            img.paste(button_img,(160,400),button_img)
        else:
            button_img = Image.open('/Users/moqu/Downloads/素材库1.0/按钮/LargeRoundedRect.png').convert("RGBA")
            button_img.resize((1600, 120))  # Button大小
            img.paste(button_img, (160, 400), button_img)

        draw.text((200, 400), buttonText, fill=(255, 255, 255), font=buttonfont)
        '''
        commodity_box = (img.size/2,img.size/2-150,300,300)
        for commodity_img in glob.glob('/Users/moqu/Downloads/素材库1.0/分类/商品1.1/'+category+'/*.png'):
            commodity_img_
            img.paste(commodity_img,commodity_box,commodity_img)
        '''

        '''
        commodity_img = Image.open('/Users/moqu/Downloads/素材库1.0/分类/商品1.1/电子数码/电子数码001.png')
        commodity_region = commodity_img
        commodity_region = commodity_region.convert("RGBA")
        commodity_region.thumbnail((300,200))
        img.paste(commodity_region,(800,200),commodity_region)
        '''

        logo_img = Image.open(r'/Users/moqu/Downloads/素材库1.0/Logo/Lazada_Group_Logo.png')
        box = (img.size[0] - 300, img.size[1] - 200,img.size[0], img.size[1])
        region = logo_img
        region = region.convert("RGBA")
        region = region.resize((box[2] - box[0], box[3] - box[1]))

        img.paste(region, box, region)

        dir = "/Users/moqu/Downloads/Outputfolder/"
        name = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
      # filename = dir + name + '_result.png'
        filename = dir + sheet.cell(row=i, column=3).value+'.png'
        sleep(1)
        img.save(filename)
        print (filename+" Done!")

        i=i+1

if __name__=='__main__':
    watermark()
